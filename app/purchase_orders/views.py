from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from typing import List

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.http import HttpRequest, HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import TemplateView
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.purchase_orders.models import PurchaseOrder, PurchaseOrderItem, Vendor
from app.purchase_orders.validators import is_smartavi


@dataclass
class PartNumber:
    name: str
    quantity: int
    description: str = None


@dataclass
class OrderRow:
    created_at: date
    order_number: str
    vendor_name: str
    part_numbers: List[PartNumber]
    eta: str = None
    supplier_ack: str = None
    comments: str = None


class UploadOrdersView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "upload.html"
    http_method_names = ["get", "post"]

    def post(self, request: HttpRequest, *args, **kwargs):
        file: InMemoryUploadedFile = request.FILES.get("file")  # type: ignore
        # TODO add validators: check file type, file size
        workbook = load_workbook(file)
        sheet: Worksheet = workbook.worksheets[0]
        try:
            part_numbers_sheet = workbook.worksheets[1]
        except IndexError:
            messages.error(request, "Not found page with part numbers")
            return HttpResponseRedirect(reverse_lazy("admin:index"))
        parsed_items = []
        vendor_names = []
        order_numbers = []
        order_part_numbers_mapping = defaultdict(list)
        for row in part_numbers_sheet.iter_rows(min_row=2, values_only=True):
            order_part_numbers_mapping[row[0]].append(
                PartNumber(name=row[1], quantity=int(row[2]), description=row[3])
            )

        for row in sheet.iter_rows(min_row=2, values_only=True):
            parsed_item = OrderRow(
                created_at=row[0],
                order_number=row[1],
                vendor_name=row[2],
                supplier_ack=row[3],
                eta=row[4],
                comments=row[5],
                part_numbers=order_part_numbers_mapping.get(row[1], []),
            )
            parsed_items.append(parsed_item)
            vendor_names.append(row[2])
            order_numbers.append(row[1])
        vendors = (
            Vendor.objects.filter(name__in=vendor_names).values("pk", "name").all()
        )
        vendors_map = {i["name"]: i["pk"] for i in vendors}
        existing_purchase_orders = set(
            PurchaseOrder.objects.filter(order_number__in=order_numbers)
            .values_list("order_number", flat=True)
            .all()
        )

        objects_created = 0
        for parsed_item in parsed_items:
            if parsed_item.order_number in existing_purchase_orders:
                continue
            obj = PurchaseOrder.objects.create(
                order_number=parsed_item.order_number,
                supplier_ack=parsed_item.supplier_ack,
                eta=parsed_item.eta,
                comments=parsed_item.comments,
                created_at=parsed_item.created_at,
                created_by_id=request.user.pk,
                vendor_id=vendors_map.get(parsed_item.vendor_name),
            )
            objects_created += 1

            for part_number_item in parsed_item.part_numbers:
                PurchaseOrderItem.objects.create(
                    part_number=part_number_item.name,
                    quantity=part_number_item.quantity,
                    description=part_number_item.description,
                    purchase_order_id=obj.pk,
                )
        if objects_created:
            messages.success(
                request, f"Successfully loaded {objects_created} purchase orders"
            )
        return HttpResponseRedirect(reverse_lazy("admin:index"))

    def test_func(self):
        return is_smartavi(self.request.user) or self.request.user.is_superuser
